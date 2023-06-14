from sklearn.feature_extraction.text import CountVectorizer
import pandas as pd
import string
import re
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
import pickle
from openpyxl import load_workbook

# Read the data file
# Change the path to the current file and save the row number of the feedback we need to predict
path = "Thesis/De La Salle University – Dasmariñas Student Feedback(1-187) - Copy.xlsx"
dataset = pd.read_excel(path)  # pathing to the file
# Save the total numbers of respondents in a constant
len_dataset = len(dataset)
# Remove N/A sentiments and repeated responses
dataset = dataset[["Feedback", "Comment/Suggestion for Improvement", "Sentiment"]]
lemmatizer = WordNetLemmatizer()
dataset = dataset.dropna(axis=0)


# Cleaning dataset and tokenize
def clean_text(txt):
    txt_nopunct = "".join([x for x in txt if x not in string.punctuation])  # Remove punctuation
    txt_nopunct = txt_nopunct.lower()  # Lowercase all
    txt_token = re.split("\W", txt_nopunct)  # Tokenization
    txt_stop = ([x for x in txt_token if x not in stopwords.words("english")])  # Remove stop words
    txt_stop = [x for x in txt_stop if x != ""]
    txt_lemmatize = []
    for word in txt_stop:
        txt_lemmatize.append(lemmatizer.lemmatize(word))  # Lemmatize each word
    txt_final = list(set(txt_lemmatize))  # Remove redundant words in one statement
    return txt_final


saved_model = pickle.load(open("Thesis/StackedModel.pkl", "rb"))
vectorizer_words = CountVectorizer(analyzer=clean_text, min_df=8)
vectored_content = vectorizer_words.fit_transform(dataset["Feedback"])


# Get the text from Excel that needs prediction and its row number
# ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
# Replace "text to be predicted with the variable"
text_to_be_predicted = "example text"
new_content = vectorizer_words.transform([text_to_be_predicted])
model_prediction = int(saved_model.predict(new_content))

# Save the model_prediction to the row,col of the sentiment
excel_editor = load_workbook(path)
sheet = excel_editor.active

for ws in excel_editor.worksheets:
    for index, row in enumerate(ws.iter_rows(), start=1):
        if ws.cell(row=index, column=11).value is None:
            ws.cell(row=index, column=11).value = model_prediction

excel_editor.save(path)

# from sklearn.feature_extraction.text import CountVectorizer
# import pandas as pd
# import string
# import re
# from nltk.corpus import stopwords
# from nltk.stem import WordNetLemmatizer

# #read the data file
# #change the path to the current file and save the row number of the feedback we need to predict
# #path = file path
# path = "Thesis\De La Salle University – Dasmariñas Student Feedback(1-187) - Copy.xlsx"
# dataset = pd.read_excel(path)  # pathing to the file
# #save the total numbers of respondent in a constant
# len_dataset = len(dataset)
# #remove N/A sentiments and repeated responses
# dataset = dataset[["Feedback","Comment/Suggestion for Improvement","Sentiment"]]
# lemmatizer = WordNetLemmatizer()
# dataset = dataset.dropna(axis=0)
# #cleaning dataset and tokenize
# def clean_text(txt):
#     txt_nopunct = "".join([x for x in txt if x not in string.punctuation]) #remove punctuation
#     txt_nopunct = txt_nopunct.lower() #lower casing all
#     txt_token = re.split("\W", txt_nopunct) # tokenization
#     txt_stop = ([x for x in txt_token if x not in stopwords.words("english")]) #remove stop words
#     txt_stop = [x for x in txt_stop if x!=""] 
#     txt_lemmatize = []      
#     for word in txt_stop:
#         txt_lemmatize.append(lemmatizer.lemmatize(word)) #lemmatizing each word
#     txt_final = list(set(txt_lemmatize)) #removing redundunt words in one statement
#     return txt_final

# import pickle
# saved_model = pickle.load(open("Thesis\StackedModel.pkl","rb"))
# vectorizer_words = CountVectorizer(analyzer=clean_text, min_df=8)
# vectored_content = vectorizer_words.fit_transform(dataset["Feedback"])


# #get the text from excel that needs prediction and its row number
# #^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
# #replace "text to be predicted with the variable"
# new_content = vectorizer_words.transform([f"{text to be predicted}"])
# model_prediction = int(saved_model.predict(new_content))

# #save the model_prediction to the row,col of the sentiment
# from openpyxl import load_workbook
# excel_editor = load_workbook(path)
# sheet = excel_editor.active

# for ws in excel_editor.worksheets:
#     for index, row in enumerate(ws.rows, start=1):
#         if ws.cell(row=index, column=11).value is None:
#             # ws.cell(row=##row of the sentiment##, column=11).value = model_prediction

# excel_editor.save(path)
            